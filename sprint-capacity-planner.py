# library modules
import datetime
import time
import openpyxl
import sys
from dataclasses import fields
import pandas as pd
import numpy as np
import xlrd
# from decimal import Decimal

# local include files
import credentials as cr
import config as cfg
from scp_classes import Default, BankHoliday, Sprint, EmployeeVacation, Employee
from scp_mapping import VACATION_TYPE_SICK_LEAVE, EMPLOYEE_STATUS_TEXT
from scp_mapping import EMPLOYEE_AVAILABLE, EMPLOYEE_ON_VACATION, EMPLOYEE_ON_SICK_LEAVE, EMPLOYEE_NOT_ON_PROJECT, \
    EMPLOYEE_ON_WEEKEND, EMPLOYEE_ON_BANK_HOLIDAY

start_time = time.time()


def load_to_dataframe(file_details):
    full_file_name = file_details['path'] + '/' + file_details['filename']
    skip_rows = 0
    try:
        df = pd.read_excel(full_file_name, sheet_name=file_details['sheet'], na_filter=False)
        for i in range(len(df)):
            if df.iloc[i, 0] == "":
                skip_rows += 1
            else:
                if skip_rows > 0:
                    skip_rows += 1
                break
        df = pd.read_excel(full_file_name, sheet_name=file_details['sheet'], na_filter=False, skiprows=skip_rows)
        loaded_columns = df.columns.values
        expected_columns = np.array(file_details['columns'])
        if not np.array_equal(loaded_columns, expected_columns):
            df = "Sheet Process Error: " + file_details[
                'sheet'] + ": loaded columns " + str(loaded_columns) + " doesn't much to expected columns " + str(
                file_details['columns'])
    except FileNotFoundError:
        df = "File Open Error: File Not Found: " + full_file_name
    except Exception as e:
        df = "File Open Error: " + e.__str__() + " File: " + full_file_name
    return df


def load_to_dataframe_with_exit_on_error(file_details):
    df = load_to_dataframe(file_details)
    if not isinstance(df, pd.DataFrame):
        # print error message and exit
        print(df)
        exit(1)
    return df


def load_to_dictionary(file_details, load_max_col, load_max_row=50):
    xlsfile = file_details['path'] + '/' + file_details['filename']
    sheet_name = file_details['sheet']
    all_rows = []
    if load_max_col < 1:
        all_rows[0] = 'load_max_col must be greater than 0'
        return all_rows
    if load_max_row < 1:
        all_rows[0] = 'load_max_row must be greater than 0'
        return all_rows
    try:
        wb = openpyxl.load_workbook(xlsfile, read_only=True, data_only=True)
    except Exception:
        print("Unexpected error:", sys.exc_info()[0])
        return all_rows
    keys = []
    try:
        sheet = wb[sheet_name]
    except Exception:
        print("Unexpected error:", sys.exc_info()[0])
        return all_rows
    for first_row in sheet.iter_rows(min_row=1, max_row=1, min_col=0, max_col=load_max_col, values_only=True):
        for key in first_row:
            keys.append(key)
    for row in sheet.iter_rows(min_row=2, max_row=load_max_row, min_col=0, max_col=load_max_col, values_only=True):
        if row[0] is None:
            continue
        one_row = {}
        for i in range(0, load_max_col):
            one_row[keys[i]] = row[i]
        all_rows.append(one_row)
    return all_rows


def load_to_objects(object_class, full_file_name, sheet_name, load_max_col, load_max_row=50, ):
    xlsfile = full_file_name['path'] + '/' + full_file_name['filename']
    all_rows = []
    if load_max_col < 1:
        all_rows[0] = 'load_max_col must be greater than 0'
        return all_rows
    if load_max_row < 1:
        all_rows[0] = 'load_max_row must be greater than 0'
        return all_rows
    try:
        wb = openpyxl.load_workbook(xlsfile, read_only=True, data_only=True)
    except Exception:
        print("Unexpected error:", sys.exc_info()[0])
        return all_rows
    try:
        sheet = wb[sheet_name]
    except Exception:
        print("Unexpected error:", sys.exc_info()[0])
        return all_rows
    for row in sheet.iter_rows(min_row=2, max_row=load_max_row, min_col=0, max_col=load_max_col, values_only=True):
        one_object = object.__new__(object_class)
        i = 0
        for field, field_value in zip(fields(object_class), row):
            one_object.__setattr__(field.name, field_value)
            i = i + 1

        all_rows.append(one_object)

    return all_rows


def vacations_data_process(list_of_vacations, list_of_bank_holidays, list_of_employees):
    list_of_employee_leaves = []
    list_of_vacations = sorted(list_of_vacations, key=lambda x: (x['EMPLOYEE'], x['START DATE']))
    list_of_bank_holidays = sorted(list_of_bank_holidays, key=lambda x: (x['Date']))
    list_of_employees = sorted(list_of_employees, key=lambda x: (x['Name']))
    for employee in list_of_employees:
        employee_bank_holidays = get_employee_bank_holidays(employee['Name'], list_of_employees, list_of_bank_holidays)
        employee_obj = EmployeeVacation(employee['Name'], employee_bank_holidays, employee['FTE'])
        filtered_vacations = [vac for vac in list_of_vacations if vac['EMPLOYEE'] == employee['Name']]
        for vacation_record in filtered_vacations:
            start_date_date = vacation_record['START DATE']
            if isinstance(start_date_date, datetime.datetime):
                start_date_date = start_date_date.date()
            end_date_date = vacation_record['END DATE']
            if isinstance(end_date_date, datetime.datetime):
                end_date_date = end_date_date.date()
            employee_obj.add_vacations_from_range(start_date_date, end_date_date)
        list_of_employee_leaves.append(employee_obj)
    return list_of_employee_leaves


def bank_holidays_data_process(list_of_bank_holidays):
    all_holidays = []
    countries = []
    list_of_bank_holidays = sorted(list_of_bank_holidays, key=lambda x: (x['Date']))
    current_date = ""
    current_date_date = ""
    for item in list_of_bank_holidays:
        if item['Date'] != current_date:
            if current_date != "":
                all_holidays.append({'date': current_date_date, 'countries': countries})
                countries = []
            current_date = item['Date']
            if isinstance(current_date, datetime.datetime):
                current_date_date = current_date.date()
            else:
                current_date_date = current_date
        countries.append(item['Country'])
    all_holidays.append({'date': current_date_date, 'countries': countries})
    holidays_obj = BankHoliday(all_holidays)
    return holidays_obj


def get_employee_bank_holidays(employee, list_of_employees, list_of_bank_holidays):
    employee_bank_holidays = []
    employee_dict = next((item for item in list_of_employees if item['Name'] == employee), None)
    if type(employee_dict) == dict:
        country_code = employee_dict['Country']
        for bank_holiday in list_of_bank_holidays:
            if bank_holiday['Country'] == country_code:
                employee_bank_holidays.append(bank_holiday['Date'].date())

    return employee_bank_holidays


def sprints_data_process(list_of_sprints, list_of_employee_vacations):
    list_of_sprint_details = []
    for sprint in list_of_sprints:
        sprint_details_obj = Sprint(sprint['Sprint ID'], sprint['Start date'], sprint['End date'])
        date_range = pd.bdate_range(sprint['Start date'], sprint['End date'])
        for single_date in date_range:
            for vacation_obj in list_of_employee_vacations:
                if vacation_obj.is_on_holiday(single_date):
                    sprint_details_obj.add_employee_on_leave(single_date, vacation_obj.name, vacation_obj.fte)
                else:
                    sprint_details_obj.add_employee_available(single_date, vacation_obj.name, vacation_obj.fte)
        list_of_sprint_details.append(sprint_details_obj)
    return list_of_sprint_details


def validate_employees_list(list_of_employees):
    data_error = False
    for employee in list_of_employees:
        if employee['Start date on project'] is None:
            print('ERROR!', employee['Name'], ': start date cannot be empty!')
            data_error = True
        if employee['End date on project'] is not None and (
                employee['End date on project'] < employee['Start date on project']):
            print('ERROR!', employee['Name'], ': end date(', employee['End date on project'].date(),
                  ') cannot be earlier than start date (', employee['Start date on project'].date(), ')!')
            data_error = True
    return data_error


def test4():
    emp1 = Employee('Bela', 'HU', 0.8, datetime.datetime(2020, 1, 15, 9, 0), datetime.datetime(2020, 7, 31, 17, 00))
    emp1.add_vacation(datetime.datetime(2020, 7, 25, 10, 11))
    emp1.add_sick_leave(datetime.datetime(2020, 6, 25, 10, 11))
    emp1.add_vacation(datetime.datetime(2020, 6, 24, 10, 11))
    emp1.add_vacation(datetime.datetime(2020, 8, 3, 10, 11))
    emp1.add_extra_working_day(datetime.date(2020, 6, 13))
    emp1.add_extra_working_day(datetime.date(2020, 6, 27))
    emp1.add_extra_working_day(datetime.date(2020, 6, 20))
    x_date = datetime.date(2020, 6, 23)
    print(x_date, emp1.is_available(x_date), emp1.status(x_date))
    print(emp1.get_nominal_fte_in_date_range(datetime.date(2020, 1, 15), datetime.date(2020, 1, 28)))
    print(emp1.get_nominal_fte_in_date_range(datetime.date(2020, 1, 14), datetime.date(2020, 1, 27)))
    print(emp1.get_nominal_fte_in_date_range(datetime.date(2020, 1, 2), datetime.date(2020, 1, 15)))
    print(emp1.get_nominal_fte_in_date_range(datetime.date(2020, 7, 31), datetime.date(2020, 8, 13)))


def test5(list_of_employee_obj):
    x_date = datetime.date(2020, 12, 12)
    for item in list_of_employee_obj:
        print(item.name, x_date, 'available:', item.is_available(x_date), 'status:',
              EMPLOYEE_STATUS_TEXT[item.status(x_date)])


def test6(list_of_sprint_obj):
    for item in list_of_sprint_obj:
        print(item.sprint, item.start_date.date(), item.end_date.date(), 'FTE available:',
              item.get_total_fte_available(), 'FTE on leave:', item.get_total_fte_on_leave(),
              'Capacity:', item.get_sprint_capacity(),
              'Dev team size:', item.get_dev_team_size_total())
        # if item.sprint == "Sprint 55":
        #     for member in item.members_available:
        #         print(member)
        #     for fte_date in item.fte_available:
        #         print(fte_date)


def employee_data_process(list_of_employees, list_of_vacations, list_of_bank_holidays, list_of_extra_sick_leaves,
                          list_of_extra_working_days):
    list_of_employee_objects = []
    for employee in list_of_employees:
        employee_obj = Employee(employee['Name'], employee['Country'], employee['FTE'],
                                employee['Start date on project'], employee['End date on project'])
        filtered_extra_working_days = [extra_working_day for extra_working_day in list_of_extra_working_days if
                                       extra_working_day['Country'] == employee_obj.country]
        for extra_working_day_record in filtered_extra_working_days:
            employee_obj.add_extra_working_day(extra_working_day_record['Date'])
        filtered_vacations = [vac for vac in list_of_vacations if vac['EMPLOYEE'] == employee['Name']]
        for vacation_record in filtered_vacations:
            start_date_date = vacation_record['START DATE']
            if isinstance(start_date_date, datetime.datetime):
                start_date_date = start_date_date.date()
            end_date_date = vacation_record['END DATE']
            if isinstance(end_date_date, datetime.datetime):
                end_date_date = end_date_date.date()
            if vacation_record['VACATION TYPE'] == VACATION_TYPE_SICK_LEAVE:
                employee_obj.add_sick_leaves_from_range(start_date_date, end_date_date)
            else:
                employee_obj.add_vacations_from_range(start_date_date, end_date_date)
        filtered_sick_leaves = [leave for leave in list_of_extra_sick_leaves if leave['Name'] == employee['Name']]
        for sick_leave_record in filtered_sick_leaves:
            start_date_date = sick_leave_record['Start date']
            if isinstance(start_date_date, datetime.datetime):
                start_date_date = start_date_date.date()
            end_date_date = sick_leave_record['End date']
            if isinstance(end_date_date, datetime.datetime):
                end_date_date = end_date_date.date()
            employee_obj.add_sick_leaves_from_range(start_date_date, end_date_date)
        filtered_bank_holidays = [bank_holiday for bank_holiday in list_of_bank_holidays if
                                  bank_holiday['Country'] == employee['Country']]
        for bank_holiday_record in filtered_bank_holidays:
            employee_obj.add_bank_holiday(bank_holiday_record['Date'])
        list_of_employee_objects.append(employee_obj)
    return list_of_employee_objects


def sprint_and_employee_data_process(list_of_sprints, list_of_employee_obj):
    list_of_sprint_obj = []
    for sprint in list_of_sprints:
        sprint_obj = Sprint(sprint['Sprint ID'], sprint['Start date'], sprint['End date'])
        date_range = pd.date_range(sprint_obj.start_date, sprint_obj.end_date)
        for employee_obj in list_of_employee_obj:
            if employee_obj.country == 'HU':
                sprint_obj.add_to_dev_team_size_hu(
                    employee_obj.get_nominal_fte_in_date_range(sprint_obj.start_date, sprint_obj.end_date))
            elif employee_obj.country == 'UK':
                sprint_obj.add_to_dev_team_size_uk(
                    employee_obj.get_nominal_fte_in_date_range(sprint_obj.start_date, sprint_obj.end_date))

            for single_date in date_range:
                if employee_obj.is_available(single_date):
                    sprint_obj.add_employee_available(single_date, employee_obj.name, employee_obj.fte)
                elif employee_obj.status(single_date) in (
                        EMPLOYEE_ON_VACATION, EMPLOYEE_ON_SICK_LEAVE, EMPLOYEE_ON_BANK_HOLIDAY):
                    sprint_obj.add_employee_on_leave(single_date, employee_obj.name, employee_obj.fte)
        list_of_sprint_obj.append(sprint_obj)
    return list_of_sprint_obj


def write_overview_sheet(list_of_sprint_obj, file_details):
    pass


def main():
    now = datetime.datetime.now()
    print("Started at:", now.strftime("%Y-%m-%d %H:%M:%S"))
    print("#####----------------------------------#####")
    # bank_holidays_list = load_to_dictionary(cfg.bank_holidays_excel, 3, 100)
    # vacations_list = load_to_dictionary(cfg.vacations_excel, 16, 200)
    # developers_list = load_to_dictionary(cfg.team_members_excel, 5)
    # sprints_list = load_to_dictionary(cfg.sprints_excel, 7, 100)
    # defaults_list = load_to_dictionary(cfg.defaults_excel, 2)
    # extra_sick_leaves_list = load_to_dictionary(cfg.extra_sick_leaves_excel, 3)
    # extra_working_days_list = load_to_dictionary(cfg.extra_working_days_excel, 3)
    # if validate_employees_list(developers_list):
    #     exit(-1)
    # bank_holidays_obj = bank_holidays_data_process(bank_holidays_list)
    # employee_obj_list = employee_data_process(developers_list, vacations_list, bank_holidays_list,
    #                                           extra_sick_leaves_list, extra_working_days_list)
    # sprint_obj_list = sprint_and_employee_data_process(sprints_list, employee_obj_list)
    # write_overview_sheet(sprint_obj_list, cfg.overview_excel)
    # test6(sprint_obj_list)

    bank_holidays_df = load_to_dataframe_with_exit_on_error(cfg.bank_holidays_excel)
    print(bank_holidays_df.head(100))
    team_members_df = load_to_dataframe_with_exit_on_error(cfg.team_members_excel)
    print(team_members_df.head(100))
    sprints_df = load_to_dataframe_with_exit_on_error(cfg.sprints_excel)
    print(sprints_df.head(100))
    extra_sick_leaves_df = load_to_dataframe_with_exit_on_error(cfg.extra_sick_leaves_excel)
    print(extra_sick_leaves_df.head(100))
    extra_working_days_df = load_to_dataframe_with_exit_on_error(cfg.extra_working_days_excel)
    print(extra_working_days_df.head(100))
    extra_working_days_df = load_to_dataframe_with_exit_on_error(cfg.extra_working_days_excel)
    print(extra_working_days_df.head(100))
    vacation_requests_df = load_to_dataframe_with_exit_on_error(cfg.vacation_requests_excel)
    print(vacation_requests_df.head(100))

    # print(bank_holidays_obj)
    # print(bank_holidays_obj.is_holiday(datetime.date(2020, 12, 25)))

    # sp1 = sprints[0]
    # sp1.define_workday_dates()
    print("#####----------------------------------#####")
    print("--- Execution time: %s seconds ---" % (time.time() - start_time))


if __name__ == '__main__':
    main()
